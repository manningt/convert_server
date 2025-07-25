#!/usr/bin/env python3
'''
2 functions:
1. make_guests_per_caller_lists(in_filename) -> Caller_lists
2. make_caller_pdfs(caller_mapping_dict, guest_dict, date_str)

The input file is an Excel file with 3 sheets: 'guest-to-caller', 'callers', 'guests'
See class definition below for the NamedTuple Caller_lists returned by make_guests_per_caller_lists(in_filename)

The output is a PDF file for each caller with a list of guests for the next Friday.

'''

import sys, os
try:
   from openpyxl import load_workbook
   from fpdf import FPDF
except Exception as e:
   print(e)
   sys.exit(1)

import argparse
from pathlib import Path
from flask import current_app

from typing import NamedTuple  #not to be confused with namedtuple in collections
class Caller_lists(NamedTuple):
    success: bool = False
    message: str = ''
    caller_mapping_dict: dict = {}
    guest_dict: dict = {}
    invalid_usernames: list = []
    invalid_caller_names: list = []
    no_guest_list: list = []
    guests_without_caller: list = []

def remove_unicode(string):
   try:
      ascii_str = string.encode(encoding="ascii", errors="strict")
      return string
   except:
      # convert some specific unicode to ascii, then replace rest with spaces
      ascii_str = string.replace("’", "'") #right single quotation mark: \u2019	&rsquo; (I've)
      ascii_str = ascii_str.replace("‘", "'")  #left single quotation mark: \u2018	&lsquo;
      ascii_str = ascii_str.replace("“",'"') #left double quote \u201C
      ascii_str = ascii_str.replace("”",'"') #right double quote \u201D
      ascii_str = ascii_str.replace("…",'...') #horizontal ellipsis \u2026 &hellip;
      ascii_str = ascii_str.replace("–", "-") # ndash: \u2013
      ascii_byte_str = ascii_str.encode(encoding="ascii", errors="ignore")
      return ascii_byte_str.decode("utf-8")


def make_guests_per_caller_lists(in_filename):
   # returns the tuple Caller_lists
   GUESTS_SHEET_NAME = 'Master List'
   GUEST_CALLER = 0
   GUEST_FIRSTNAME = 1
   GUEST_LASTNAME = 2
   GUEST_USERNAME = 3
   GUEST_PASSWORD = 4
   # the following 3 columns were hidden previously to 2025-07-24, and are not used in the reports:
   # GUEST_HOUSEHOLD_ID = 5
   # GUEST_DROP_LOCATION = 6
   # GUEST_ADDRESS = 7
   GUEST_TOWN = GUEST_PASSWORD +1
   GUEST_PHONE = GUEST_TOWN + 1
   GUEST_NOTES = GUEST_PHONE + 1

   Caller_lists()
   Caller_lists.success = False # default value didn't seem to work

   try:
      workbook = load_workbook(in_filename, data_only=True)
   except Exception as e:
      Caller_lists.message = f"Could not read file '{os.path.basename(in_filename)}': {e}"
      return Caller_lists
   
   sheetnames = workbook.sheetnames
   # print(f"{workbook.sheetnames=}")

   # check that expected sheets are in the excel spreadsheet (other sheets are ignored)
   EXPECTED_SHEETNAMES = {GUESTS_SHEET_NAME}
   for name in EXPECTED_SHEETNAMES:
      if name not in sheetnames:
         Caller_lists.message = f"Expected sheet '{name}' not found in file '{os.path.basename(in_filename)}'"
         return Caller_lists
   
   guests_with_no_caller = []
   callers_with_no_guest_list = []

   # make a dictionary of guest data to be used for generating reports.
   mapping_dict = {}
   guest_dict = {}
   invalid_usernames = []
   invalid_caller_names = []
   # min_row=2 to skip header row
   for row in workbook[GUESTS_SHEET_NAME].iter_rows(min_row=2, max_row=300, \
                                                   min_col=1, max_col=(GUEST_NOTES+1), values_only=True):

      if row[GUEST_USERNAME] is None and row[GUEST_LASTNAME] is None:
         continue #skip blank rows
      else:
         # remove characters above the font range, e.g. "’", "“"
         cleaned_values = [""] * (GUEST_NOTES+1)
         for index in [GUEST_CALLER, GUEST_USERNAME, GUEST_FIRSTNAME, GUEST_LASTNAME, GUEST_PASSWORD, GUEST_TOWN, GUEST_PHONE, GUEST_NOTES]:
            if row[index] is not None and isinstance(row[index], str):
               cleaned_values[index] = remove_unicode(row[index])
         # separate caller from the caller notes:
         if cleaned_values[GUEST_CALLER] is not None:
            if cleaned_values[GUEST_CALLER].find(" ") > 0:
               caller,caller_note = cleaned_values[GUEST_CALLER].split(" ",1)
            else:
               caller = cleaned_values[GUEST_CALLER]
               caller_note = "" 
         # print(f"{caller= } {caller_note=}")
         if len(caller) < 3:
            invalid_caller_names.append(f"{caller} (calling {cleaned_values[GUEST_FIRSTNAME]} {cleaned_values[GUEST_LASTNAME]})")
         elif len(cleaned_values[GUEST_USERNAME]) < 3:
            invalid_usernames.append(f"{cleaned_values[GUEST_FIRSTNAME]} {cleaned_values[GUEST_LASTNAME]}")
         else:
            guest_dict[cleaned_values[GUEST_USERNAME]]= {'caller_note':caller_note,'First':cleaned_values[GUEST_FIRSTNAME], 
            'Last':cleaned_values[GUEST_LASTNAME], 'Username': cleaned_values[GUEST_USERNAME],
            'PW':cleaned_values[GUEST_PASSWORD],'Town':cleaned_values[GUEST_TOWN], 'Phone':cleaned_values[GUEST_PHONE], 
            'Notes':cleaned_values[GUEST_NOTES]}
            if caller in mapping_dict:
               mapping_dict[caller].append(cleaned_values[GUEST_USERNAME])
            else:
               mapping_dict[caller] = [cleaned_values[GUEST_USERNAME]]
   '''
   Caller_lists.guest_dict={'PConnors': {'Caller Note': '', 'First': 'Patrick', 'Last': 'Connors', 'Username': 'PConnors', 'PW': 'Foodpantry1', 'Town': 'Newburyport', 'Phone': '(978) 255-4252', 'Notes': 'Delivery. Typically available in morning'},
   
   Caller_lists.caller_mapping_dict= {'BarbaraK': ['PConnors', 'Rsmith2'], 'BarbaraP': ['LAnderson1', 'HillC', 'DMcCarthy', 'WeckbacherA'], 'Caitlin': ['JBashaw', 'DHardy'],
   '''
   
   Caller_lists.caller_mapping_dict = mapping_dict
   Caller_lists.message = "No error"  # put in to _tuplegetter(1, 'Alias for field number 1' error
   Caller_lists.guest_dict = guest_dict
   Caller_lists.no_guest_list = callers_with_no_guest_list
   Caller_lists.invalid_usernames = invalid_usernames
   Caller_lists.invalid_caller_names = invalid_caller_names
   Caller_lists.guests_without_caller = guests_with_no_caller
   Caller_lists.success = True

   return Caller_lists


def filter_callers(caller_mapping_dict):
   filtered_dict = {}
   # see if the 'Change' row has a value, or if the caller does not equal the normal_caller
   callers_with_substitutes = []
   for caller, guests in caller_mapping_dict.items():
      changed = False
      for caller_guest_dict in guests:
         if caller_guest_dict['change'] is not None:
            changed = True
            # print(f"{caller}-{caller_guest_dict['guest']} {caller_guest_dict['change']=}")
         if caller != caller_guest_dict['normal_caller']:
            changed = True
            callers_with_substitutes.append(caller_guest_dict['normal_caller'])
            # print(f"{caller}-{caller_guest_dict['guest']} {caller_guest_dict['normal_caller']=}")
      if changed:
         filtered_dict[caller] = guests
         # print(f"filtered_dict_length={len(filtered_dict)}")

   # add callers_with_substitutes:
   for caller in callers_with_substitutes:
      filtered_dict[caller] = caller_mapping_dict[caller]
      
   return filtered_dict


def make_caller_pdfs(caller_mapping_dict, guest_dict, date_str, out_pdf_dir='.'):
   # PDF writing examples:
   #  https://medium.com/@mahijain9211/creating-a-python-class-for-generating-pdf-tables-from-a-pandas-dataframe-using-fpdf2-c0eb4b88355c
   #  https://py-pdf.github.io/fpdf2/Tutorial.html
   success_list = []
   failure_list = []
   print_count = 0
   for caller, guests in caller_mapping_dict.items():
      # print(f'{caller=} {guests=}')
      try:
         pdf = FPDF(orientation="L", format="letter") # default units are mm; adding , unit="in" inserts blank pages
         pdf.add_page()
         pdf.set_font("Helvetica", size=14)
         pantry_day = date_str.replace("_"," ")
         pdf.cell(0, 10, f"{caller} - Pantry Day: {pantry_day}", align="C")
         pdf.ln(10)

         with pdf.table(col_widths=(14,11,13,14,13,13,14,30), line_height=6) as table:
            pdf.set_font(size=12)
            header = ['Caller notes', 'First', 'Last', 'UserName', 'Password', 'Town', 'Phone', 'Notes about guest']
            row = table.row()
            for column in header:
               row.cell(column)

            for this_guest_username in guests:
               if print_count > 0:
                  print(f'{this_guest_username=}')
                  print_count -= 1
               if this_guest_username in guest_dict:
                  this_guest_dict = guest_dict[this_guest_username]
               else:
                  print(f'{this_guest_username=} is not in guest_dict')
                  continue
               row_data = [this_guest_dict['caller_note'], this_guest_dict['First'], this_guest_dict['Last'], \
                           this_guest_username, \
                           this_guest_dict['PW'], this_guest_dict['Town'], this_guest_dict['Phone'], \
                           this_guest_dict['Notes']]
               # print(f"{row_data=}")
               row = table.row()
               for item in row_data:
                  row.cell(str(item), v_align="T")

         pdf.output(os.path.join(out_pdf_dir, f'{caller.replace(" ","_")}_{date_str}.pdf'))
         success_list.append(caller)
      except Exception as e:
         try:
            current_app.logger.warning(f"PDF for {caller} failed: {e}")
         except:
            print(f"PDF for {caller} failed: {e}")
         failure_list.append(caller)
         sys.exit(1)  # stop on first failure
   return (success_list, failure_list)


def get_fridays_date_string():
   import datetime
   today = datetime.date.today()
   # print(f"{today.weekday()=}")
   # Sunday is 6, Monday is 0, Tuesday is 1, Wednesday is 2, Thursday is 3, Friday is 4, Saturday is 5
   # find the next Friday
   days_ahead = 4 - today.weekday()
   if days_ahead <= 0: # Target day already happened this week
      days_ahead += 7
   # print(f"{days_ahead=}")
   target_date = today + datetime.timedelta(days=days_ahead)
   return target_date.strftime('%Y-%m-%d')


if __name__ == "__main__":
   argParser = argparse.ArgumentParser()
   argParser.add_argument("input", type=str, help="input filename with path")

   args = argParser.parse_args()

   if args.input is None:
      sys.exit("No file selected to parse.")
   elif Path(args.input).is_file():
      filename = args.input
   else:
      sys.exit("file selected is not a file.")

   Caller_lists = make_guests_per_caller_lists(filename)
   if not Caller_lists.success:
      print(f"Failure: {Caller_lists.message}")
      sys.exit(1)

   print(f"Callers with no guests: {Caller_lists.no_guest_list}")

   date_str = get_fridays_date_string()
   # filter callers are callers with substitutes
   # filtered_callers_dict = filter_callers(Caller_lists.caller_mapping_dict)
   # make_caller_pdfs(filtered_callers_dict, Caller_lists.guest_dict, date_str, out_pdf_dir="/tmp")
   make_caller_pdfs(Caller_lists.caller_mapping_dict, Caller_lists.guest_dict, date_str, out_pdf_dir="/tmp")

'''
The following function is not used in the current code, but is kept here for reference.
It was used to make a mapping of guests to callers using a sheet named guest-to-caller, and to return a dictionary of guests

def make_guests_per_caller_lists_w_mapping(in_filename):
   # returns the tuple Caller_lists
   GUESTS_SHEET_NAME = 'Master List'
   GUEST_FIRSTNAME = 1
   GUEST_LASTNAME = 2
   GUEST_USERNAME = 3
   GUEST_PASSWORD = 4
   # in old versions of the sheet before 2025-07-24, there were 3 hidden columns (5,6,7) that are not used in the PDFs
   GUEST_TOWN = GUEST_PASSWORD + 1
   GUEST_PHONE = GUEST_TOWN + 1
   GUEST_NOTES = GUEST_PHONE + 1

   Caller_lists()
   Caller_lists.success = False # default value didn't seem to work

   try:
      workbook = load_workbook(in_filename, data_only=True)
   except Exception as e:
      Caller_lists.message = f"Could not read file '{os.path.basename(in_filename)}': {e}"
      return Caller_lists
   
   sheetnames = workbook.sheetnames
   # print(f"{workbook.sheetnames=}")

   # check that expected sheets are in the excel spreadsheet (other sheets are ignored)
   EXPECTED_SHEETNAMES = {'guest-to-caller', 'callers', GUESTS_SHEET_NAME}
   for name in EXPECTED_SHEETNAMES:
      if name not in sheetnames:
         Caller_lists.message = f"Expected sheet '{name}' not found in file '{os.path.basename(in_filename)}'"
         return Caller_lists

   #make dictionary of caller, [guests]
   mapping_dict = {}
   is_header = True
   for row in workbook['callers'].rows:
      # skip header row; there is only 1 column: the list of callers
      if is_header:
         is_header = False
      elif row[0].value is not None:
         mapping_dict[row[0].value] = []
      else:
         continue
   
   guests_with_no_caller = []
   is_header = True
   for row in workbook['guest-to-caller'].rows:
      # columns: 0=Guest, 1=Caller, 2=Note, 3=Normal Caller
      if is_header:
         is_header = False
      elif row[0].value is not None:
         if row[1].value is not None:
            # print(f'{row[0].value} -> {row[1].value}')
            # a list containing the guest name and note to the caller's list:
            if row[2].value is not None and isinstance(row[2].value, str):
               caller_note = remove_unicode(row[2].value)
            else:
               caller_note = ""
            key = row[1].value
            if key in mapping_dict:
               mapping_dict[key].append({'guest':row[0].value, 'caller_note':caller_note, 'normal_caller':row[3].value})
            else:
               mapping_dict[key] = [{'guest':row[0].value, 'caller_note':caller_note, 'normal_caller':row[3].value}]
         else:
            guests_with_no_caller.append(row[0].value)
            current_app.logger.info(f"guest '{row[0].value}' does not have a caller")
      else:
         continue

   # current_app.logger.info(f"{mapping_dict=}")
   # mapping_dict={'Barb': [{'guest': 'LAnderson1', 'caller_note': '', 'normal_caller': 'Barb'}, {'guest': 'DMcCarthy', 'caller_note': '', 'normal_caller': 'Barb'}

   callers_with_no_guest_list = []
   for caller, guests in mapping_dict.items():
      if len(guests) == 0:
         callers_with_no_guest_list.append(caller)
   # current_app.logger.info(f"{callers_with_no_guest_list= }") 

   # remove caller with no guests from mapping_dict
   for caller in callers_with_no_guest_list:   
         mapping_dict.pop(caller)

   # make a dictionary of guest data to be used for generating reports.
   guest_dict = {}
   is_header = True
   invalid_usernames = []
   for row in workbook[GUESTS_SHEET_NAME].rows:
      if is_header:
         is_header = False
      elif row[GUEST_USERNAME].value is None and row[GUEST_LASTNAME].value is None:
         continue #skip blank rows
      else:
         # remove characters above the font range, e.g. "’", "“"
         cleaned_values = [""] * (GUEST_NOTES+1)
         for index in [GUEST_USERNAME, GUEST_FIRSTNAME, GUEST_LASTNAME, GUEST_PASSWORD, GUEST_TOWN, GUEST_PHONE, GUEST_NOTES]:
            if row[index].value is not None and isinstance(row[index].value, str):
               cleaned_values[index] = remove_unicode(row[index].value)
         if len(cleaned_values[GUEST_USERNAME]) < 3:
            invalid_usernames.append(f"{cleaned_values[GUEST_FIRSTNAME]} {cleaned_values[GUEST_LASTNAME]}")
         else:
            guest_dict[cleaned_values[GUEST_USERNAME]]= {'First':cleaned_values[GUEST_FIRSTNAME], 'Last':cleaned_values[GUEST_LASTNAME], 
            'PW':cleaned_values[GUEST_PASSWORD],'Town':cleaned_values[GUEST_TOWN], 'Phone':cleaned_values[GUEST_PHONE], 
            'Notes':cleaned_values[GUEST_NOTES]}      
   # print(f"{guest_dict=}")
   # guest_dict={'Guest1': {'First': 'Guest', 'Last': 1.0, 'PW': 'secret', 'Town': 'Newbury', 'Phone': '978.555.0000', 'Notes': 'call early'}, 'Guest2': {'First': 'Guest', 'Last': 2.0, 'PW': 'secret', 'Town': 'Newbury', 'Phone': '978.555.0000', 'Notes': 'call 3 times'}, 'Guest3': {'First': 'Guest', 'Last': 3.0, 'PW': 'secret', 'Town': 'Newbury', 'Phone': '978.555.0000', 'Notes': 'call late'}}
   
   Caller_lists.caller_mapping_dict = mapping_dict
   Caller_lists.guest_dict = guest_dict
   Caller_lists.no_guest_list = callers_with_no_guest_list
   Caller_lists.invalid_usernames = invalid_usernames
   Caller_lists.guests_without_caller = guests_with_no_caller
   Caller_lists.success = True

   return Caller_lists
'''